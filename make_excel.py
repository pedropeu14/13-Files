#!/usr/bin/env python3
"""Gera 13F_analise.xlsx a partir de data/analysis.json."""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
d = json.load(open(f"{HERE}/data/analysis.json", encoding="utf-8"))

F = "Arial"
H_FILL = PatternFill("solid", start_color="1F3864")
H_FONT = Font(name=F, bold=True, color="FFFFFF", size=10)
B_FONT = Font(name=F, size=10)
BOLD = Font(name=F, bold=True, size=10)
GREEN = Font(name=F, size=10, color="006100")
RED = Font(name=F, size=10, color="9C0006")
THIN = Border(bottom=Side(style="thin", color="D9D9D9"))
MM = '#,##0.0'
INT = '#,##0'

wb = Workbook()

def sheet(ws, headers, widths):
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, c, h); cell.font = H_FONT; cell.fill = H_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"

def row(ws, r, vals, fmts=None, font=None):
    for c, v in enumerate(vals, 1):
        cell = ws.cell(r, c, v); cell.font = font or B_FONT; cell.border = THIN
        if fmts and fmts.get(c): cell.number_format = fmts[c]

slug2name = {m["slug"]: m["manager"] for m in d["managers"]}
short = {s: n.split("(")[0].strip() for s, n in slug2name.items()}

# ---- 1. Resumo -------------------------------------------------------------
ws = wb.active; ws.title = "Resumo"
sheet(ws, ["Gestor", "CIK", "Trimestres com 13F", "Último trimestre",
           "Nº posições", "Valor 13F (US$ mm)"], [42, 12, 22, 15, 11, 16])
r = 2
for m in sorted(d["managers"], key=lambda x: -x["latest"]["total"]):
    row(ws, r, [m["manager"], m["cik"], len(m["periods"]), m["latest"]["period"],
                m["latest"]["n"], m["latest"]["total"] / 1e6],
        {5: INT, 6: MM}); r += 1
row(ws, r, ["TOTAL", "", "", "", f"=SUM(E2:E{r-1})", f"=SUM(F2:F{r-1})"],
    {5: INT, 6: MM}, BOLD)
ws.cell(r + 2, 1, "Fonte: SEC EDGAR, formulários 13F-HR. Gerado em " + d["generated"]
        + ". Valores conforme reportado (ver aba Notas).").font = Font(name=F, italic=True, size=9)

# ---- 2. Consenso -----------------------------------------------------------
ws = wb.create_sheet("Consenso")
sheet(ws, ["De", "Para", "Papel", "Put/Call", "CUSIP", "Abriram (nº)", "Fecharam (nº)",
           "Aumentaram (nº)", "Reduziram (nº)", "Balanço (net)", "Quem abriu",
           "Quem fechou", "Valor aberto (US$ mm)"],
      [11, 11, 34, 9, 12, 12, 12, 13, 12, 12, 34, 34, 16])
r = 2
for tr in sorted(d["consensus"], reverse=True):
    a, b = tr.split("__")
    for x in d["consensus"][tr]:
        npart = len(x["opened"]) + len(x["closed"]) + len(x["increased"]) + len(x["decreased"])
        if npart < 2:
            continue
        net = x["net_interest"]
        row(ws, r, [a, b, x["issuer"], x["put_call"], x["cusip"],
                    len(x["opened"]), len(x["closed"]), len(x["increased"]),
                    len(x["decreased"]), net,
                    ", ".join(short.get(s, s) for s in x["opened"]),
                    ", ".join(short.get(s, s) for s in x["closed"]),
                    x["value_open"] / 1e6],
            {13: MM}, GREEN if net > 0 else (RED if net < 0 else None))
        r += 1
ws.auto_filter.ref = f"A1:M{r-1}"

# ---- 3. Entradas e Saídas ---------------------------------------------------
ws = wb.create_sheet("Entradas e Saídas")
sheet(ws, ["Gestor", "De", "Para", "Movimento", "Papel", "Put/Call", "CUSIP",
           "Valor (US$ mm)", "Ações"], [30, 11, 11, 10, 34, 9, 12, 14, 14])
r = 2
for slug in sorted(d["changes"]):
    for t in d["changes"][slug]:
        for p in t["new"]:
            row(ws, r, [short.get(slug, slug), t["from"], t["to"], "ENTRADA",
                        p["issuer"], p["put_call"], p["cusip"], p["value"] / 1e6,
                        p["shares"]], {8: MM, 9: INT}, GREEN); r += 1
        for p in t["exited"]:
            row(ws, r, [short.get(slug, slug), t["from"], t["to"], "SAÍDA",
                        p["issuer"], p["put_call"], p["cusip"], p["value"] / 1e6,
                        p["shares"]], {8: MM, 9: INT}, RED); r += 1
ws.auto_filter.ref = f"A1:I{r-1}"

# ---- 4. Aumentos e Reduções -------------------------------------------------
ws = wb.create_sheet("Aumentos e Reduções")
sheet(ws, ["Gestor", "De", "Para", "Movimento", "Papel", "Put/Call", "CUSIP",
           "Δ ações", "Δ %", "Valor atual (US$ mm)"],
      [30, 11, 11, 11, 34, 9, 12, 14, 9, 16])
r = 2
for slug in sorted(d["changes"]):
    for t in d["changes"][slug]:
        for kind, tag, fnt in (("increased", "AUMENTO", GREEN), ("decreased", "REDUÇÃO", RED)):
            for p in t[kind]:
                row(ws, r, [short.get(slug, slug), t["from"], t["to"], tag,
                            p["issuer"], p["put_call"], p["cusip"], p["shares_delta"],
                            (p["pct"] or 0) / 100, p["value"] / 1e6],
                    {8: INT, 9: '0.0%', 10: MM}, fnt); r += 1
ws.auto_filter.ref = f"A1:J{r-1}"

# ---- 5. Carteiras atuais ----------------------------------------------------
ws = wb.create_sheet("Carteiras (último tri)")
sheet(ws, ["Gestor", "Trimestre", "Papel", "Classe", "Put/Call", "CUSIP",
           "Valor (US$ mm)", "Ações", "% da carteira"],
      [30, 11, 34, 14, 9, 12, 14, 14, 12])
r = 2
for m in sorted(d["managers"], key=lambda x: -x["latest"]["total"]):
    slug, p = m["slug"], m["latest"]["period"]
    for cusip, issuer, cls, pc, value, shares in d["holdings"][slug][p]:
        row(ws, r, [short.get(slug, slug), p, issuer, cls, pc, cusip,
                    value / 1e6, shares,
                    f'=G{r}/SUMIFS($G:$G,$A:$A,A{r})'],
            {7: MM, 8: INT, 9: '0.0%'}); r += 1
ws.auto_filter.ref = f"A1:I{r-1}"

# ---- 6. Notas ----------------------------------------------------------------
ws = wb.create_sheet("Notas")
ws.column_dimensions["A"].width = 120
notes = [
    "13F Search — análise das carteiras de grandes gestores (SEC EDGAR, formulários 13F-HR).",
    f"Gerado em {d['generated']}. Trimestres cobertos: {', '.join(d['periods'])}.",
    "",
] + d["notes"] + [
    "",
    "Limitações do 13F: mostra apenas posições LONG em ações/ADRs/opções listadas nos EUA;",
    "não inclui posições vendidas (short), bonds, moedas ou posições internacionais fora dos EUA.",
    "Divulgado até 45 dias após o fim do trimestre — as posições podem já ter mudado.",
    "Duquesne (Druckenmiller) não havia arquivado o 13F de 2026-03-31 até a data de geração.",
    "Scion (Burry) deixou de reportar após 2025-09-30. Greenlight (Einhorn) sem 13F desde 2023.",
]
for i, n in enumerate(notes, 1):
    ws.cell(i, 1, n).font = Font(name=F, size=10, bold=(i == 1))

wb.save(f"{HERE}/13F_analise.xlsx")
print("13F_analise.xlsx salvo")
